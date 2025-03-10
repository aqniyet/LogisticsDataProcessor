import os
import pandas as pd
from openpyxl import load_workbook
import logging
from typing import List, Dict, Tuple, Optional

from app.database.operations import get_active_routes, get_matrix_mappings
from app.utils.file_utils import ensure_directory_exists

logger = logging.getLogger(__name__)

class ExpenseProcessor:
    """
    Process expense files using RouteID data and apply 1C mappings.
    """
    
    def __init__(self, config):
        """Initialize the ExpenseProcessor with configuration."""
        self.config = config
        self.base_directory = config.get('base_directory', './data')
        
        # Setup output directory
        self.output_directory = os.path.join(self.base_directory, 'Expenses_processed')
        
        # Ensure directory exists
        ensure_directory_exists(self.output_directory)
    
    def format_value(self, value) -> str:
        """Format values consistently to ensure proper matching."""
        try:
            if pd.notna(value):
                if isinstance(value, float) and value.is_integer():
                    return str(int(value)).zfill(8)
                elif str(value).isdigit():
                    return str(int(value)).zfill(8)
            return str(value).strip()
        except Exception:
            return str(value).strip()
    
    def find_and_clean_headers(self, df, expected_columns) -> Tuple[pd.DataFrame, int]:
        """Find headers by inspecting the top rows and remove rows before the header."""
        for i in range(min(10, len(df))):
            header_row = df.iloc[i].astype(str).str.strip().str.lower()
            if all(col.lower() in header_row.values for col in expected_columns):
                df.columns = df.iloc[i]
                df = df.drop(df.index[:i + 1]).reset_index(drop=True)
                return df, i + 1
        raise ValueError("Expected headers not found in the file.")
    
    def process_expense_file(self, file_path: str, reference_data: pd.DataFrame, 
                           active_values: set, matrix_mappings: pd.DataFrame,
                           output_folder: str) -> Tuple[bool, str]:
        """Process a single expense file with route ID data and 1C mappings."""
        try:
            # Load the workbook
            original_wb = load_workbook(file_path)
            original_ws = original_wb.active
            
            # Read the data
            raw_data = pd.read_excel(file_path, sheet_name=0, header=None)
            main_data, header_row_index = self.find_and_clean_headers(
                raw_data, ['номер вагона', 'номер документа']
            )
            
            if main_data.empty:
                return False, "Empty data after header detection"
            
            # Standardize column names
            main_data.rename(columns={
                'Номер вагона': 'Вагон №',
                'Номер документа': 'Накладная №'
            }, inplace=True)
            
            # Format values for consistent matching
            main_data['Вагон №'] = main_data['Вагон №'].apply(self.format_value)
            main_data['Накладная №'] = main_data['Накладная №'].apply(self.format_value)
            
            # Merge with reference data
            merged_data = main_data.merge(
                reference_data, how='left', on=['Вагон №', 'Накладная №']
            )
            
            # Ensure ЗНП column is numeric and fill nulls with 0
            output_column_name = 'ЗНП'
            merged_data[output_column_name] = pd.to_numeric(
                merged_data.get(output_column_name, 0), errors='coerce'
            ).fillna(0).astype(int)
            
            # Add or update ЗНП column in Excel
            if output_column_name not in [cell.value for cell in original_ws[header_row_index]]:
                original_ws.cell(
                    row=header_row_index, 
                    column=original_ws.max_column + 1, 
                    value=output_column_name
                )
            
            znp_col_index = [cell.value for cell in original_ws[header_row_index]].index(output_column_name) + 1
            
            # Write ЗНП values to Excel
            for row_idx, value in enumerate(merged_data[output_column_name], start=header_row_index + 1):
                original_ws.cell(row=row_idx, column=znp_col_index, value=value)
            
            # Add 1C column using matrix mappings
            merged_data['для 1С'] = merged_data['ЗНП'].apply(
                lambda x: self.find_in_matrix_and_check(x, matrix_mappings, active_values)
            )
            
            # Add or update для 1С column in Excel
            if 'для 1С' not in [cell.value for cell in original_ws[header_row_index]]:
                original_ws.cell(
                    row=header_row_index, 
                    column=original_ws.max_column + 1, 
                    value='для 1С'
                )
            
            dla_1c_col_index = [cell.value for cell in original_ws[header_row_index]].index('для 1С') + 1
            
            # Write для 1С values to Excel
            for row_idx, value in enumerate(merged_data['для 1С'], start=header_row_index + 1):
                original_ws.cell(row=row_idx, column=dla_1c_col_index, value=value)
            
            # Save the processed file
            output_file_path = os.path.join(output_folder, os.path.basename(file_path))
            original_wb.save(output_file_path)
            
            return True, output_file_path
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {str(e)}")
            return False, str(e)
    
    def find_in_matrix_and_check(self, value, matrix_mappings, active_values):
        """Find a mapping in the matrix that leads to an active value."""
        # Convert value to string and strip whitespace
        value_str = str(value).strip()
        
        # If the value is already active, return it
        if value_str in active_values:
            return value_str
        
        # Find all mappings where this value is the source
        possible_mappings = matrix_mappings[matrix_mappings['source_value'] == value_str]
        
        # Check each target value to see if it's active
        for _, mapping in possible_mappings.iterrows():
            target = mapping['target_value']
            if target in active_values:
                return target
            
            # If target isn't active, check if it has any further mappings
            secondary_mappings = matrix_mappings[matrix_mappings['source_value'] == target]
            for _, sec_mapping in secondary_mappings.iterrows():
                if sec_mapping['target_value'] in active_values:
                    return sec_mapping['target_value']
        
        return "value is not active"
    
    def process_expense_folder(self, expense_folder: str, route_id_data_path: str) -> Dict:
        """Process all expense files in a folder."""
        # Load reference data
        reference_data = pd.read_csv(route_id_data_path, encoding='utf-8')
        reference_data['Вагон №'] = reference_data['Вагон №'].apply(self.format_value)
        reference_data['Накладная №'] = reference_data['Накладная №'].apply(self.format_value)
        
        # Get active routes and matrix mappings
        active_routes = get_active_routes()
        active_values = set(active_routes['route_id'].astype(str))
        matrix_mappings = get_matrix_mappings()
        
        # Initialize counters
        processed_files = 0
        skipped_files = 0
        error_files = []
        
        # Process all Excel files in the input folder
        for file in os.listdir(expense_folder):
            if file.lower().endswith(('.xlsx', '.xls')):
                file_path = os.path.join(expense_folder, file)
                
                # Process file with both RouteID and 1C mappings in one step
                success, result = self.process_expense_file(
                    file_path, reference_data, active_values, matrix_mappings, self.output_directory
                )
                
                if success:
                    processed_files += 1
                else:
                    skipped_files += 1
                    error_files.append((file, result))
        
        # Return processing summary
        return {
            "processed_files": processed_files,
            "skipped_files": skipped_files,
            "error_files": error_files
        }